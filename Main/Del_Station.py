import os
import sys

import tkinter as tk
from tkinter import ttk

import openpyxl
from openpyxl import *
from openpyxl.styles import *

from Main import station_data # Import to be able to read station_data variable

class Del_Station():
    def __init__(self):
        self.Remove_Station()

    def Remove_Station(self):
        wb = openpyxl.load_workbook(station_data)
        ws = wb.active
        rows = ws.max_row
        cols = ws.max_column
        print(rows, cols)



