import os
import sys

import tkinter as tk
from tkinter import ttk

import openpyxl
from openpyxl import *
from openpyxl.styles import *

from Main import station_directory # Import to be able to read station_data variable

def DeleteStation(self):
    wb = openpyxl.load_workbook(station_directory)
    ws = wb.active
    rows = ws.max_row
    cols = ws.max_column
    print(rows, cols)