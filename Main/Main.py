import os
from pathlib import Path
import platform

import tkinter as tk
from tkinter import ttk

from datetime import datetime
import time

import openpyxl

from openpyxl import *
from openpyxl.styles import *

root = tk.Tk()
main_window = tk.Toplevel(root)

version = 'v0.1 06/2026'

now = datetime.now()
current_time = now.strftime("%m/%d/%y %H:%M")
print(current_time)

#### Check OS being used and change storage directory accordingly (Future use)
os_name = platform.system()

#### Create Filename
file = f'Station_Database.xlsx'

#### Set path to Home directory
home_path = Path.home()

#### Create Directories for Station Info and the Data Reports
station_info_directory = f'{home_path}\\Ryvear_Galaxy\\Station\\'
report_file_directory = f'{home_path}\\Ryvear_Galaxy\\Report\\'

#### Join Directory and Filename
station_directory = os.path.join(station_info_directory, file)
station_data = station_directory

#### Verify storage directories exist and if not create
if os.path.exists(station_info_directory): # if station information directory exists do nothing
    pass
else:
    os.makedirs(station_info_directory) # if directory does not exist create it

#### If file already exists open it. If ot create it
if os.path.isfile(station_directory):
    wb = openpyxl.load_workbook(station_directory) # open file if exists
else:
    wb = Workbook() # create file
    ws = wb.active  # Set Active Sheet
    # Creating a workbook titles
    ws['A1'] = 'Controller Type' # ESP32 type, STM32 type, or Microchip type
    ws['B1'] = 'Controller Version' # station firmware version
    ws['C1'] = 'Controller Address' # Comport, wifi, ble, or usb
    ws['D1'] = 'Controller Location' # Where is station located in the system
    ws['E1'] = 'Controller Type' # Station Version (Analog, Digital, Camera, Sensors)
    ws['F1'] = 'Station Added' # Date station added
    ws['G1'] = 'Station Removed' # Date removed maintained for history

    #### Set row 1 titles alignment, font, and style, borders
    for letter in ['A','B','C','D','E','F','G']:
        max_width = 24
        for i in letter:
            ws[f'{i}1'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{i}1'].font = Font(name = 'Arial', bold = True, italic = False, size = 12, color = "000000") # Change font and color
            ws[f'{i}1'].border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin')) # Top, Left, Right, Bottom, Thick, thin, dotted, dashed
            ws.column_dimensions[f'{letter}'].width = 28
        for row in range(ws.max_row):
            ws[f'{i}1'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{i}1'].font = Font(name = 'Arial', bold = True, italic = False, size = 8, color = "000000") # Change font and color
            ws[f'{i}1'].border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin')) # Top, Left, Right, Bottom, Thick, thin, dotted, dashed

    wb.save(station_directory) # Save data to file

main_window = tk.Toplevel(root)
datetime_label = tk.Label(main_window, text=current_time, font=("Arial", 8), width=20, height=1)
datetime_label.place(x=175, y=10)
version_label = tk.Label(main_window, text=version, font=("Arial", 8), width=20, height=1)
version_label.place(x=0, y=10)

main_window.title("Ryvear Main Page")
system_color = main_window.cget('bg')
main_window.configure(background=system_color)
main_window.geometry("300x50+50+50")

menubar = tk.Menu(main_window)
station_menu = tk.Menu(menubar, tearoff=0)
station_menu.add_command(label = "Add Station", command = lambda: InstallStation())
station_menu.add_command(label = "Remove Station", command = lambda: DeleteStation())
menubar.add_cascade(label = "Station", menu = station_menu)

data_capture_menu = tk.Menu(menubar, tearoff=0)
data_capture_menu.add_command(label = "Select Station", command = lambda: DataCapture(1))
data_capture_menu.add_command(label = "Select By Type", command = lambda: DataCapture(2))
data_capture_menu.add_command(label = "Select By Location", command = lambda: DataCapture(3))
menubar.add_cascade(label = "Data Capture", menu = data_capture_menu)

data_review_menu = tk.Menu(menubar, tearoff=0)
data_review_menu.add_command(label = "Station History", command = lambda: DataStore())
data_review_menu.add_command(label = "History By Type", command = lambda: DataStore())
data_review_menu.add_command(label = "History By Location", command = lambda: DataStore())
menubar.add_cascade(label = "Data Review", menu = data_review_menu)

help_menu = tk.Menu(menubar, tearoff=0)
help_menu.add_command(label = "Local Help Document", command = lambda: HelpLocal())
help_menu.add_command(label = "Online Help Document", command = lambda: HelpOnline())
help_menu.add_command(label = "Tech Support", command = lambda: TechSupport())
help_menu.add_command(label="Version Information", command = lambda: SysVersion())
help_menu.add_command(label="Contact Information", command = lambda: Contact())
menubar.add_cascade(label = "Help", menu = help_menu)

quit_menu = tk.Menu(menubar, tearoff=0)
quit_menu.add_command(label = "Quit", command = exit)
menubar.add_cascade(label = "Quit", menu = quit_menu)
main_window.config(menu = menubar)

def InstallStation():
    var = tk.IntVar()
    main_window.geometry("400x500+50+50")
    datetime_label.place(x = 275, y = 475)
    version_label.place(x = 0, y = 475)
    def on_select(event):
        selected = CommChoiceCombobox.get()
        if selected == 'Serial_Comm':
            SerialComm()
        elif selected == 'WiFi':
            Wifi_Search()
        elif selected == 'Bluetooth':
            Bluetooth_Search()
        elif selected == 'USB':
            USB_Search()
        else:
            print("invalid Selection")
    CommSelect = ['', 'Serial_Comm', 'WiFi', 'Bluetooth', 'USB']
    CommChoiceCombobox = ttk.Combobox(main_window, values = CommSelect, state="readonly")
    CommChoiceCombobox.pack(pady = 25, padx = 0)
    CommChoiceCombobox.current(0)
    CommChoiceCombobox.bind("<<ComboboxSelected>>", on_select)

def SerialComm():
    import serial
    from serial.tools import list_ports
    ports = serial.tools.list_ports.comports()
    print('Looking For New Stations On Serial Ports')
    comportnum = []
    for port in ports: # uncomment
        ser = serial.Serial('com7', baudrate = 115200, timeout = 2)
        try:
            if ser.isOpen():
                comportnum.append(port.device)
                comportnum.append(port.description)
                comportnum.append(datetime.now().date())
                comportnum.append(datetime.now().time())
                station_comm = port.description.split() # uncomment
                if station_comm[1] == 'CH340':
                    ser.write(b'Send Station Status')
                    ser.flush()
                    time.sleep(1) # indent for if()
                    rcvd_data = ser.read(5) # indent for if()
                    print(rcvd_data)
                    xfer_size = int(rcvd_data, 16)
                    print(xfer_size)
                    ser.flush()
                    for i in range (xfer_size):
                        rcvd_data = ser.read(2048)  # indent for if()
                        print(rcvd_data)
                        char_data_size = int.from_bytes(rcvd_data, "big")
                        print(char_data_size)
                        time.sleep(1)  # indent for if()
                        response_data = ser.read(char_data_size)  # indent for if()
                        print(response_data) # indent for if()
                        ser.flush()
        except serial.SerialException as e: # If error connecting move on to next port
            pass

def Wifi_Search():
    print('Looking For New Stations On Wifi')

def Bluetooth_Search():
    print('Looking For New Stations On Bluetooth')

def USB_Search():
    print('Looking For New Stations On USB')

def DeleteStation():
    rows = ws.max_row
    cols = ws.max_column
    print(rows, cols)

def DataCapture(data_type):
    if data_type == 1:
        print('Data Capture Single Station')
    elif data_type == 2:
        print('Data Capture Multiple Stations By Type')
    else:
        print('Data Capture Multiple Stations By Location')

def DataStore():
    print('DataStore')

def HelpLocal():
    print('HelpLocal')

def HelpOnline():
    print('HelpOnline')

def TechSupport():
    print('TechSupport')

def SysVersion():
    print('Version')

def Contact():
    print('Contact')

def RefreshStationData():
    print('RefreshStationData')

def ViewStationData():
    print('ViewStationData')

def Exit():
    print('Exit')

main_window.mainloop()